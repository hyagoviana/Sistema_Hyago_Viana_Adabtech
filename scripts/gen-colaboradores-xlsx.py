from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "1E2044"
GOLD = "C9A227"
LIGHT = "F5F1E6"
WHITE = "FFFFFF"

wb = Workbook()

# ─── Aba Colaboradores ───────────────────────────────────────────────
ws = wb.active
ws.title = "Colaboradores"

cols = [
    ("Nome completo", 26),
    ("E-mail corporativo", 30),
    ("Telefone", 16),
    ("Cargo", 14),
    ("Time / Equipe", 18),
    ("ID ProJuris", 14),
    ("Participa da distribuição", 20),
    ("Peso (padrão 100)", 15),
    ("Ver Operacional", 15),
    ("Ver Comercial", 15),
    ("Ver Financeiro", 15),
    ("Ver Judicial", 15),
    ("Ver Controladoria", 16),
    ("Vê valores (R$)", 14),
    ("Observações", 28),
]

# Título
ws.merge_cells("A1:O1")
t = ws["A1"]
t.value = "Cadastro de Colaboradores — Sistema HV (preencher e devolver)"
t.font = Font(name="Arial", bold=True, size=13, color=WHITE)
t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 26

# Cabeçalho
thin = Side(style="thin", color="D9D2BF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for i, (name, width) in enumerate(cols, start=1):
    c = ws.cell(row=2, column=i, value=name)
    c.font = Font(name="Arial", bold=True, size=10, color=NAVY)
    c.fill = PatternFill("solid", fgColor=LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width
ws.row_dimensions[2].height = 30

# Linha de exemplo (em itálico cinza, para o Thiago apagar/copiar)
exemplo = [
    "Ana Patrícia Cruz", "anapatricia@hyagovianaadvocacia.com.br", "(82) 90000-0000",
    "Sênior", "Time TEMFC", "131021", "Sim", 100,
    "Edita", "Não vê", "Não vê", "Vê", "Não vê", "Não", "exemplo — apague esta linha",
]
for i, v in enumerate(exemplo, start=1):
    c = ws.cell(row=3, column=i, value=v)
    c.font = Font(name="Arial", size=9, italic=True, color="8A8266")
    c.alignment = Alignment(horizontal="center" if i not in (1, 2, 15) else "left", vertical="center")
    c.border = border

# Linhas em branco
for r in range(4, 26):
    for i in range(1, len(cols) + 1):
        c = ws.cell(row=r, column=i)
        c.font = Font(name="Arial", size=10)
        c.border = border
        c.alignment = Alignment(horizontal="center" if i not in (1, 2, 15) else "left", vertical="center")
    ws.cell(row=r, column=8).value = 100  # peso default

# Dropdowns (data validation)
dv_cargo = DataValidation(type="list", formula1='"Estagiário,Júnior,Sênior"', allow_blank=True)
dv_simnao = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
dv_perm = DataValidation(type="list", formula1='"Não vê,Vê,Edita"', allow_blank=True)
ws.add_data_validation(dv_cargo); dv_cargo.add("D3:D25")
ws.add_data_validation(dv_simnao); dv_simnao.add("G3:G25"); dv_simnao.add("N3:N25")
ws.add_data_validation(dv_perm); dv_perm.add("I3:M25")

ws.freeze_panes = "A3"

# ─── Aba Instruções ──────────────────────────────────────────────────
wi = wb.create_sheet("Instruções")
wi.column_dimensions["A"].width = 26
wi.column_dimensions["B"].width = 80
wi.merge_cells("A1:B1")
h = wi["A1"]
h.value = "Instruções de preenchimento"
h.font = Font(name="Arial", bold=True, size=13, color=WHITE)
h.fill = PatternFill("solid", fgColor=NAVY)
h.alignment = Alignment(horizontal="left", vertical="center")
wi.row_dimensions[1].height = 26

instr = [
    ("Nome completo", "Nome do colaborador como deve aparecer no sistema."),
    ("E-mail corporativo", "E-mail @hyagovianaadvocacia — é por ele que a pessoa recebe o convite e faz login."),
    ("Telefone", "Opcional."),
    ("Cargo", "Estagiário, Júnior ou Sênior (menu suspenso). Só o SÊNIOR participa do rodízio de distribuição; o time aparece na agenda junto."),
    ("Time / Equipe", "Nome do time a que a pessoa pertence (ex.: Time TEMFC)."),
    ("ID ProJuris", "Identificador do usuário no ProJuris (o código interno). Essencial para o motor — se não tiver, deixe em branco que a gente busca."),
    ("Participa da distribuição", "Sim/Não. Em geral: Sênior = Sim; Júnior/Estagiário = Não."),
    ("Peso (padrão 100)", "Todo mundo começa com 100 (distribui igual). Reduza (ex.: 50) para quem está saindo receber menos; aumente para quem entrou receber mais."),
    ("Ver Operacional/Comercial/Financeiro/Judicial/Controladoria", "Para cada módulo: 'Não vê', 'Vê' (só leitura) ou 'Edita'. Assim eu já configuro as permissões de cada um."),
    ("Vê valores (R$)", "Sim/Não — se a pessoa pode enxergar valores financeiros."),
    ("Observações", "Qualquer detalhe (ex.: exceção de responsável exclusivo, férias etc.)."),
]
r = 3
for k, v in instr:
    a = wi.cell(row=r, column=1, value=k)
    a.font = Font(name="Arial", bold=True, size=10, color=NAVY)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = wi.cell(row=r, column=2, value=v)
    b.font = Font(name="Arial", size=10)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    wi.row_dimensions[r].height = 30
    r += 1

import os
out = os.path.join(os.path.dirname(__file__), "..", "docs", "reunioes", "Cadastro-Colaboradores-Sistema-HV.xlsx")
out = os.path.abspath(out)
wb.save(out)
print("OK:", out)
