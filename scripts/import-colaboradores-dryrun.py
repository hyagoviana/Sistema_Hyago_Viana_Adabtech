#!/usr/bin/env python3
# M15 (2026-08-08) — DRY-RUN do importador de colaboradores.
# Lê docs/reunioes/Cadastro-Colaboradores-PREENCHIDO-2026-08-08.xlsx, normaliza
# (trim \r\n), classifica ativo x arquivado (M17), aplica as regras do motor
# (M8/M9/M13) e EMITE: (1) um relatório humano no stdout e (2) um JSON
# normalizado em scripts/colaboradores-normalizado.json que o importador TS
# (--apply) vai consumir. NÃO escreve no banco, NÃO manda e-mail.
#
# Uso:  python scripts/import-colaboradores-dryrun.py

import json
import os
import sys

from openpyxl import load_workbook

XLSX = "docs/reunioes/Cadastro-Colaboradores-PREENCHIDO-2026-08-08.xlsx"
OUT = "scripts/colaboradores-normalizado.json"

# eligible_complex = só 4 (áudio Thiago; Bruno=Maxwel, Hudson=Wdyson CONFIRMADOS).
COMPLEX_NAMES = ["maxwel", "wdyson", "ana patr", "keilane"]

PERFIL_ROLE = {
    "administrador": "admin",
    "coordenador": "operacional",
    "usuario padrao": "operacional",
    "financeiro": "financeiro",
}
CARGO_MAP = {
    "senior": "senior",
    "junior": "junior",
    "estagiario": "estagiario",
    "prestador de servico": "prestador_servico",
    "administrador": "administrador",
}
ACCESS_MAP = {"nao ve": "none", "ve": "view", "edita": "edit"}


def s(v):
    """str + trim de \r\n e espaços; None -> ''."""
    if v is None:
        return ""
    return str(v).replace("\r", " ").replace("\n", " ").strip()


def deacc(v):
    """lower + remove acentos (p/ casar chaves)."""
    import unicodedata

    return "".join(
        c for c in unicodedata.normalize("NFD", s(v).lower()) if unicodedata.category(c) != "Mn"
    ).strip()


def yn(v):
    return deacc(v) == "sim"


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    if not os.path.exists(XLSX):
        print(f"ERRO: planilha nao encontrada: {XLSX}")
        sys.exit(1)

    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Colaboradores"] if "Colaboradores" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    # índice por nome de coluna (deacc)
    col = {deacc(h): i for i, h in enumerate(header) if h}

    def g(row, name):
        i = col.get(deacc(name))
        return row[i] if i is not None and i < len(row) else None

    people = []
    warnings = []
    for row in rows[1:]:
        nome = s(g(row, "Nome completo"))
        if not nome:
            continue  # linha vazia
        email = s(g(row, "E-mail corporativo")).lower()
        perfil = deacc(g(row, "Perfil"))
        cargo_raw = deacc(g(row, "Cargo"))
        status_colab = deacc(g(row, "Status colaborador/e-mail"))
        status_pj = deacc(g(row, "Status usuário projuris"))
        pes = s(g(row, "ID ProJuris"))
        pes_valido = pes.upper().startswith("PES.")
        peticionante = yn(g(row, "Peticionante"))
        participa = yn(g(row, "Participa da distribuição padrão"))
        peso_raw = g(row, "Peso (padrão 100)")
        try:
            peso = int(float(peso_raw)) if peso_raw not in (None, "") else 100
        except (TypeError, ValueError):
            peso = 100
        time_eq = s(g(row, "Time / Equipe"))
        obs = s(g(row, "Observações"))

        # Classificação ativo x arquivado (M17).
        arquivado = (
            status_pj == "desabilitado"
            or status_colab in ("inativo",)
            or not email
        )

        role = PERFIL_ROLE.get(perfil, "operacional")
        cargo = CARGO_MAP.get(cargo_raw)
        eligible_complex = any(k in deacc(nome) for k in COMPLEX_NAMES)

        # Permissões por módulo (5) + vê valores.
        perms = {}
        for mod, colname in [
            ("operacional", "Ver Operacional"),
            ("comercial", "Ver Comercial"),
            ("financeiro", "Ver Financeiro"),
            ("judicial", "Ver Judicial"),
            ("controladoria", "Ver Controladoria"),
        ]:
            perms[mod] = ACCESS_MAP.get(deacc(g(row, colname)), None)
        can_view_values = yn(g(row, "Vê valores (R$)"))

        # Avisos de consistência.
        if not arquivado and participa and cargo in ("junior", "estagiario"):
            warnings.append(f"{nome}: participa=Sim mas cargo={cargo} (regra: só sênior na fila geral)")
        if not arquivado and not pes_valido:
            warnings.append(f"{nome}: sem ID ProJuris válido (PES.*) — mapping fica pendente")
        if not arquivado and participa and not peticionante:
            warnings.append(f"{nome}: participa=Sim mas peticionante=Não (contraditório)")

        people.append(
            {
                "nome": nome,
                "email": email or None,
                "role": role,
                "perfil": perfil or None,
                "cargo": cargo,
                "time": None if time_eq in ("", "-") else time_eq,
                "status_projuris": "desabilitado" if status_pj == "desabilitado" else ("habilitado" if status_pj == "habilitado" else None),
                "arquivado": arquivado,
                "projuris_id": pes if pes_valido else None,
                "peticionante": peticionante,
                "participa_distribuicao_padrao": participa,
                "peso": peso,
                "eligible_complex": eligible_complex,
                "perms": perms,
                "can_view_values": can_view_values,
                "obs": obs or None,
            }
        )

    ativos = [p for p in people if not p["arquivado"]]
    arquivados = [p for p in people if p["arquivado"]]
    complexos = [p["nome"] for p in people if p["eligible_complex"]]
    fila_geral = [p["nome"] for p in ativos if p["participa_distribuicao_padrao"] and p["peticionante"]]

    # JSON normalizado p/ o importador TS.
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({"people": people}, f, ensure_ascii=False, indent=2)

    # Relatório humano.
    print("=" * 72)
    print("DRY-RUN — Importação de colaboradores (M15). NADA foi escrito no banco.")
    print("=" * 72)
    print(f"Total de linhas úteis: {len(people)}")
    print(f"  • ATIVOS (viriam como convite/registro ativo): {len(ativos)}")
    print(f"  • ARQUIVADOS (registro SEM acesso, sem e-mail): {len(arquivados)}")
    print()
    print(f"Fila GERAL (peticionante=Sim E participa=Sim): {fila_geral or '—'}")
    print(f"  (esperado pela planilha: Keilane, Maxwel, Wdyson)")
    print()
    print(f"eligible_complex=true (4 esperados): {complexos or '—'}")
    print()
    print(f"AVISOS ({len(warnings)}):")
    for w in warnings:
        print(f"  [!] {w}")
    print()
    print("— ATIVOS —")
    for p in ativos:
        print(
            f"  {p['nome']:<32} | {p['email'] or '(sem e-mail)':<42} | role={p['role']:<11} "
            f"cargo={p['cargo']} time={p['time']} PES={p['projuris_id']} "
            f"petic={'S' if p['peticionante'] else 'N'} part={'S' if p['participa_distribuicao_padrao'] else 'N'} "
            f"peso={p['peso']} complex={'S' if p['eligible_complex'] else 'N'} vê$={'S' if p['can_view_values'] else 'N'}"
        )
    print()
    print("— ARQUIVADOS (sem acesso) —")
    for p in arquivados:
        print(f"  {p['nome']:<32} | {p['email'] or '(sem e-mail)':<42} | PES={p['projuris_id']} status_pj={p['status_projuris']}")
    print()
    print(f"JSON normalizado salvo em: {OUT}")


if __name__ == "__main__":
    main()
